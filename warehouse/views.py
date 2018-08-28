from django.shortcuts import render
from .forms import StockInForm
from .models import StockIn
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import load_workbook
from shop.models import Product

# Create your views here.
@login_required
def Import(request):
    if request.method == 'POST':
        form = StockInForm(request.POST)
        if form.is_valid():
            stock_in = form.save(commit=False)
            user = User.objects.get(username=request.user.username)
            stock_in.create_by = user
            stock_in.save()
            messages.success(request, 'Bạn đã nhập {} sản phẩm {} vào kho thành công.'.format(stock_in.quantity, stock_in.product))
        form = StockInForm()
        return render(request, 'warehouse/input.html', {'form': form})
    else:
        form = StockInForm()
    context = {
        'form': form,
    }
    return render(request, 'warehouse/input.html', context)

@login_required
def import_list(request):
    import_list = StockIn.objects.all()
    title = ''
    search = request.GET.get('q')
    if search:
        import_list = import_list.filter(Q(product__name__contains=search.lower()) | Q(product__name__contains=search.upper()) | Q(create_by__username=search))
        title ='TÌM KIẾM THEO {}'.format(search.upper())
    paginator = Paginator(import_list, 15)
    page = request.GET.get('page')
    try:
        imports = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer deliver the first page
        imports = paginator.page(1)
    except EmptyPage:
        # If page is out of range deliver last page of results
        imports = paginator.page(paginator.num_pages)
    context = {
        'imports': imports,
        'title': title,
    }
    return render(request, 'warehouse/import_list.html', context)

@login_required
def load_excel(request):
    if request.method == 'GET':
        return render(request, 'warehouse/import_by_excel.html', {'disabled': "disabled=disabled",})
    elif request.method == 'POST':
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size

        wb = load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        firstline = True
        for row in worksheet.iter_rows():
            if firstline:  # skip first line
                firstline = False
                continue
            row_data = list()
            for cell in row:
                if cell.value:
                    row_data.append(str(cell.value))
            if row_data != []:
                excel_data.append(row_data)
        context = {
                'excel_data': excel_data,
                'disabled': "disabled=enabled",
            }
        return render(request, 'warehouse/import_by_excel.html', context)

def import_by_excel(request):
    if request.method == 'POST':
        product_code_array = request.POST.getlist('p_code_array[]')
        origin_price_array = request.POST.getlist('o_price_array[]')
        sale_price_array = request.POST.getlist('s_price_array[]')
        quantity_array = request.POST.getlist('quantity_array[]')
        action = request.POST.get('action')
        user = User.objects.get(username=request.user.username)
        if product_code_array and action:
            try:
                if action == "import":
                    for product_code in product_code_array:
                        try:
                            StockIn.objects.create(product=Product.objects.get(product_code=product_code),
                                                   origin_price=origin_price_array[product_code_array.index(product_code)],
                                                   sale_price=sale_price_array[product_code_array.index(product_code)],
                                                   quantity=int(quantity_array[product_code_array.index(product_code)]),
                                                   create_by=user)
                        except Exception as e:
                            print(e)
                            pass
                return JsonResponse({'status': 'ok'})
            except:
                pass
    return JsonResponse({'status': 'ko'})